# import pandas as pd
from io import BytesIO
import openpyxl #ediciones
from core.models import Extraccion
from dataextraction.argentina_extraccion import ExtraccionArgentina
from collections import defaultdict
import locale
from core.models import VencimientoImpuesto, Impuesto, Empresa, Empleado
from datetime import date, datetime
 # celdas = hoja['A1':'I13']
    # hoja.max_column
    # hoja.max_row
    # hoja.min_column
    # hoja.min_row



campos = ["periodo_fiscal", 'nombreFormulario', 'Fecha_vencimiento', 'Fecha_envio_cliente', 'Fecha_revisado', 'Revisor', 'año', 'numeroFormulario', ' n_verificacion', 'saldoPagado', 'saldoFavor', 'fecha_procesado']

class ProcesamientoExcel():

    @classmethod
    def lectura_xls(cls, file):
        # lectura_excel = openpyxl.load_workbook(filename=ruta, data_only=True)
        lectura_excel = openpyxl.load_workbook(filename=BytesIO(file.read()))
        hoja = lectura_excel.active
        lista=[]
        for fila in hoja.iter_rows(min_row=2, max_row = hoja.max_row):
            lista.append(ProcesamientoExcelTest.fila_dicc([celda.value  for celda in fila]))
        return lista

    @classmethod
    def comparo(cls, vencimiento, extrccion_pdf):
        return vencimiento.id_razonsocial == extrccion_pdf.id_razonsocial

    @classmethod
    def validar_datos(cls, dato_xls, extraccion_pdf):
        for vencimiento in dato_xls:
            if ProcesamientoExcel.comparo(vencimiento, extraccion_pdf):
                vencimiento.proceso  = VencimientoImpuesto.EstadoVencimiento.PROCESADO
            else:
                vencimiento.proceso = VencimientoImpuesto.EstadoVencimiento.NO_PROCESADO
        return dato_xls

class ProcesamientoExcelTest():

    @classmethod
    def fila_dicc(cls, fila):
        # print(type(fila[5]),fila[6], fila[7])
        excel_datos = VencimientoImpuesto(
            pais = fila[0],
            id_razonsocial = fila[1],
            nombre_empresa = fila[2],
            periodo_fiscal = fila[3],
            nombre_formulario = fila[4].strip(),
            año = fila[5].year,
            mes = fila[5].month,
            fecha_vencimiento = fila[5], # datetime.strptime(fila[5], '%d/%m/%Y'),
            fecha_entrega = fila[6], # datetime.strptime(fila[6], '%d/%m/%Y'),
            fecha_revisado = fila[7], # datetime.strptime(fila[7], '%d/%m/%Y'),
            review = fila[8],
            # proceso = fila[9],
        )

        return excel_datos

    @classmethod
    def datos_procesado(cls, rutapdf):
        ex_arg = ExtraccionArgentina()
        texto = ex_arg.lectura(ruta_archivo=rutapdf)
        lineas = ex_arg.procesamiento(texto)
        datos = ex_arg.extraccion(texto, lineas)
        return datos


if __name__ == "django.core.management.commands.shell":
    ruta = "dataextraction/Recibos/Fechas_Clientes.xlsx"
    rutapdf = "dataextraction/Recibos/ARG/AR02 BSF AR02_IVA_02.2022_F 731 DDJJ .pdf"
    datos_excel = ProcesamientoExcel.lectura_xls(ruta) #lista de diccionarios

    datos = ProcesamientoExcelTest.datos_procesado(rutapdf)
    resultado = ProcesamientoExcel.validar_datos(datos_excel, datos)
    # guardar
    for i in resultado:
        excel_datos = VencimientoImpuesto(proceso = i['proceso'])
        print(i['id_razonsocial'],i['proceso'])
