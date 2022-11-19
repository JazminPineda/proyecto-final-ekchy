
import json
from multiprocessing import context
import os
from datetime import date, datetime
from io import BytesIO

import openpyxl  # ediciones
from core.models import (Documento, Empresa, Extraccion, Pais, Proceso,
                         VencimientoImpuesto)
from dataextraction.argentina_extraccion import ExtraccionArgentina
from dataextraction.calculos import (GraficoEstadoImpuesto, GraficoEstadoMes,
                                     GraficoPeriodoImpuesto,
                                     GraficoRevisor_Estadoimpuesto)
from dataextraction.cololombia_extraccion import ExtraccionColombia
from dataextraction.lectura_excel import ProcesamientoExcel  # jaz
from dataextraction.mexico_extraccion import ExtraccionMexico
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout


def login_view(request):
    next = '/'
    if 'next' in request.GET.keys():
        next = request.GET['next']
    context = {
        "next": next,
        "title" : 'Ingresar al Sistema Ekchý'
    }
    return render(request, "login.html",context)

def autenticate(request):
    username = request.POST['email']
    password = request.POST['password']
    # next = "/"
    # if 'next' in request.GET.keys():
    next = request.GET['next']
    print(username,password,next)
    user = authenticate(request, username=username, password=password)
    if user is not None:
        login(request, user)
        return redirect(next)
    return render(request, "login.html")

def logout_api(request):
    logout(request)
    return redirect("/")

def index_view(request):
    # return HttpResponse('Hello World!')
    template = 'index.html'
    context = {
        "paragraph": 'Awesome test',
        "title" : 'Inicio'
    }

    return render(request, template, context)

@login_required
def pdf_upload_view(request):
    # return HttpResponse('Hello World!')
    empresas = Empresa.objects.all()
    template = 'pdfupload.html'
    context = {
        "empresas": empresas,
        "title" : 'Subir PDF'
    }
    return render(request, template, context)


@login_required
def pdf_upload(request):
    empresa_id = request.POST.getlist('empresa')[0]
    files = request.FILES.getlist('files')
    empresa = Empresa.objects.get(id=empresa_id)

    for file in files:
        proceso = Proceso.objects.create(estado=Proceso.Estados.INICIADO)
        proceso.save()
        proceso.refresh_from_db()
        document = Documento.objects.create(id_empresa=empresa, id_proceso=proceso, nombre=file.name, documento_pdf=file)
        document.save()

        resultado  = process_pdf(proceso=proceso, document=document, empresa=empresa)
        proceso.id_extraccion = resultado
        proceso.estado = Proceso.Estados.PROCESADO
        proceso.save()

    empresas = Empresa.objects.all()
    template = 'pdfupload.html'
    context = {
        "empresas": empresas,
        "mensaje": "Se subieron y se procesaron los pdf correctamente",
    }
    return render(request, template, context)



def get_extractor(empresa:Empresa):
    if empresa.pais == str(Pais.ARGENTINA):
        return ExtraccionArgentina()
    elif empresa.pais == str(Pais.COLOMBIA):
        return ExtraccionColombia()
    elif empresa.pais == str(Pais.MEXICO):
        return ExtraccionMexico()

def process_pdf(proceso:Proceso, document:Documento, empresa:Empresa):
    extractor = get_extractor(empresa=empresa)
    texto = extractor.lectura(os.path.join('.',document.documento_pdf.path))
    lineas = extractor.procesamiento(text=texto)
    datos = extractor.extraccion(text=texto, lineas=lineas)
    extraccion = Extraccion(
        id_razonsocial = datos[0],
        nombre_empresa = datos[1],
        periodo_fiscal = datos[2],
        año = datos[3],
        numero_formulario = datos[4],
        n_verificacion = datos[5],
        saldo_pagado = datos[6],
        saldo_favor= datos[7],
        nombre_formulario = datos[8],
        pais = empresa.pais,
        fecha_procesado = date.today(),
    )

    extraccion.save()
    extraccion.refresh_from_db()
    return extraccion


## Subida xml fecha de vencimientos
@login_required
def xml_upload_view(request):
    template = 'xml_upload.html'
    context = {
        "empresas": [],
        "title" : 'Subir Excel'
    }
    return render(request, template, context)

@login_required
def xml_upload(request):
    # empresa_id = request.POST.getlist('empresa')[0]
    files = request.FILES.getlist('files')[0]
    lista = ProcesamientoExcel.lectura_xls(files)
    procesos = Proceso.objects.all().filter(estado= Proceso.Estados.PROCESADO)
    resultados = ProcesamientoExcel.validar_datos(procesos, lista)
    vencimientos_procesados = ProcesamientoExcel.obtener_registros_vencimientos()

    for resultado in resultados:
        if f'{resultado.año}-{resultado.mes}' not in vencimientos_procesados.keys() \
                    or (f'{resultado.año}-{resultado.mes}' in vencimientos_procesados.keys() \
                        and resultado.id_razonsocial not in vencimientos_procesados[f'{resultado.año}-{resultado.mes}']):
            resultado.save()



    template = 'view_grilla_vencimientos.html'
    fecha_de_hoy = datetime.today()
    context = {
        "vencimientos": resultados,
        "fecha": fecha_de_hoy,
        "mensaje": "Se subió y se procesó el xml correctamente"
    }
    return render(request, template, context)



## Subida xml de "Pdfs periodos anteriores"
def xml_upload_period_view(request):
    template = 'xml-upload-periodos.html'
    context = {
        "carga": [],
    }
    return render(request, template, context)


def xml_upload_period(request):
    files = request.FILES.getlist('files')[0]
    lectura_excel = openpyxl.load_workbook(filename=BytesIO(files.read()))
    hoja = lectura_excel.active

    for fila in hoja.iter_rows(min_row=2, max_row = hoja.max_row):

            extraccion = VencimientoImpuesto(
                id_razonsocial = fila[0].value,
                nombre_empresa = fila[1].value,
                periodo_fiscal = fila[2].value,
                año = fila[3].value,
                mes = fila[2].value +1,
                nombre_formulario = fila[8].value,
                pais = Pais[fila[9].value.upper()],
                fecha_vencimiento = fila[10].value,
                review = fila[11].value,
                proceso = fila[12].value
            )

            extraccion.save()
    template = 'xml-upload-periodos.html'
    context = {
        "carga": []
    }
    return render(request, template, context)


## dashboard

@login_required
def dashboard_view(request):
    """Devuelve html vista """
    template = 'dashboard.html'
    context = {
        "title" : 'Dashboard'
    }

    return render(request, template, context)


def dashboard_API(request):
    """Devuelve los datos de la BD"""
    context = datos_graficos()
    vencimientos_dic = consultar_vencimientos()

    context['grafico1']['data'] = construir_datos_grafico1(vencimientos_dic)
    context['grafico2']['data'] = construir_datos_grafico2(vencimientos_dic)
    context['grafico3']['data'] = construir_datos_grafico3(vencimientos_dic)
    context['grafico4']['data'] = construir_datos_grafico4(vencimientos_dic)
    context['totalImpuestoHoy'] = calcular_total_impuestos_hoy(vencimientos_dic)
    context['impuestosProcesados'] = calcular_impuestosprocesados(vencimientos_dic)
    context['impuestosPendientes']= calcular_impuestospendientes(vencimientos_dic)

    return HttpResponse(json.dumps(context), content_type="application/json")

def calcular_total_impuestos_hoy(vencimientos):
    return len(vencimientos)

def calcular_impuestosprocesados(vencimientos):
    contador = 0
    mes_encurso = date(2022,4,4).month

    for impuesto in vencimientos:
        if impuesto['fecha_vencimiento'].month == mes_encurso and impuesto['proceso'] == 'Procesado':
            contador += 1
    return contador


def calcular_impuestospendientes(vencimientos):
    contador = 0
    mes_encurso = date(2022,4,4).month
    for impuesto in vencimientos:
        if impuesto['fecha_vencimiento'].month == mes_encurso and impuesto['proceso'] == 'Pendiente':
            contador += 1
    return contador

def consultar_vencimientos():
    vencimientos = VencimientoImpuesto.objects.all()
    vencimientos_dic = [vencimiento.__dict__ for vencimiento in vencimientos]
    return vencimientos_dic


def construir_datos_grafico1(vencimientos_dic):
    grafico1 = {'datasets': []}
    grafico_periodo = GraficoPeriodoImpuesto()
    impuestos = grafico_periodo.cantidadImpuesto(vencimientos_dic)
    imagen_uno = grafico_periodo.formato_data(impuestos)
    grafico1['datasets'].append({
        'label': 'Procesamiento de Impuestos por mes',
        'data':imagen_uno,
        'backgroundColor': [
            'rgba(54, 162, 235, 0.2)',
        ],
        'borderColor': [
            'rgba(54, 162, 235, 1)',
        ],
        'borderWidth': 1
        })

    return grafico1

def construir_datos_grafico2(vencimientos_dic):
    grafico2 = {'datasets': []}
    grafico_estado_impts = GraficoEstadoImpuesto()
    imagen_dos = grafico_estado_impts.cantidadDocumentosProcesados(vencimientos_dic)
    grafico2['datasets'].append({
        'label': 'Estado de documentos por mes',
        'data':imagen_dos,
        'backgroundColor': [
            'rgba(153, 102, 255, 0.2)',
        ],
        'borderColor': [
            'rgba(153, 102, 255, 1)',
        ],
        'borderWidth': 1
        })

    return grafico2



def construir_datos_grafico3(vencimientos_dic):
    grafico3 = {'datasets': []}
    grafico = GraficoEstadoMes()
    formato_grafico = grafico.datos_xls(vencimientos_dic)
    docts_mes = grafico.procesamiento_mes(vencimientos_dic)
    imagen_tres = grafico.union_datos(formato_grafico, docts_mes)

    imagen_tres[0]['backgroundColor']='rgba(75, 192, 192, 0.2)'
    imagen_tres[0]['borderColor']='rgba(75, 192, 192, 0.2)'
    imagen_tres[1]['backgroundColor']='rgba(255, 99, 132, 0.2)'
    imagen_tres[1]['borderColor']='rgba(255, 99, 132, 0.2)'

    grafico3['datasets'] = imagen_tres

    return grafico3


def construir_datos_grafico4(vencimientos_dic):
    grafico4 ={}
    grafico = GraficoRevisor_Estadoimpuesto()
    datos_revisor = grafico.datos_excel(vencimientos_dic)
    datos_pdf = grafico.datos_pdf(vencimientos_dic)
    union_datos = grafico.union_datos(datos_revisor,datos_pdf)
    imagen4 = grafico.formato_data(union_datos)
    grafico4['labels'] = imagen4[0]
    imagen4[1][0]['backgroundColor']= 'rgba(54, 162, 235, 0.2)'
    imagen4[1][1]['backgroundColor']= 'rgba(255, 99, 132, 0.2)'
    imagen4[1][2]['backgroundColor']= 'rgba(75, 192, 192, 0.2)'
    grafico4['datasets'] = imagen4[1]
    return grafico4


def datos_graficos():

    return {
        'grafico1':{
            'data':{
                    'datasets': [{
                        'label': 'Totales',
                        'data': {'3': 14, '2': 19, '1': 11},
                        'backgroundColor': [
                            'rgba(54, 162, 235, 0.2)',
                            # 'rgba(255, 206, 86, 0.2)',
                            # 'rgba(75, 192, 192, 0.2)',
                            # 'rgba(153, 102, 255, 0.2)',
                            # 'rgba(255, 159, 64, 0.2)'
                        ],
                        'borderColor': [

                            'rgba(54, 162, 235, 1)',
                            # 'rgba(255, 206, 86, 1)',
                            # 'rgba(75, 192, 192, 1)',
                            # 'rgba(153, 102, 255, 1)',
                            # 'rgba(255, 159, 64, 1)'
                        ],
                        'borderWidth': 1
                    }]
                }
        },
        'grafico2':{
            'data':{
                    'datasets': [{
                        'label': 'Estado de documentos por mes',
                        'data': {'OK Procesado': 26, 'Pendiente': 11, 'Corrección': 3, 'No procesado': 4},
                        'backgroundColor': [
                            'rgba(255, 99, 132, 0.2)',
                            'rgba(54, 162, 235, 0.2)',
                            'rgba(255, 206, 86, 0.2)',
                            'rgba(75, 192, 192, 0.2)',
                            'rgba(153, 102, 255, 0.2)',
                            'rgba(255, 159, 64, 0.2)'
                        ],
                        'borderColor': [
                            'rgba(255, 99, 132, 1)',
                            'rgba(54, 162, 235, 1)',
                            'rgba(255, 206, 86, 1)',
                            'rgba(75, 192, 192, 1)',
                            'rgba(153, 102, 255, 1)',
                            'rgba(255, 159, 64, 1)'
                        ],
                        'borderWidth': 1
                 }]
            }
        },
        'grafico3':{
            'data':{
                    'datasets': [
                            {
                                'label': 'OK Procesado',
                                'data': {'3': 3, '2': 16, '1': 7},
                                'backgroundColor': [
                                    'rgba(54, 162, 235, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(54, 162, 235, 1)',
                                ],
                                'borderWidth': 1

                            },
                            {
                                'label': 'Pendiente',
                                'data': {'3': 11},
                                'backgroundColor': [
                                    # 'rgba(54, 162, 235, 0.2)',
                                    'rgba(255, 99, 132, 0.2)',
                                ],
                                'borderColor': [
                                    # 'rgba(54, 162, 235, 1)',
                                    'rgba(255, 99, 132, 1)',
                                ],
                                'borderWidth': 1
                                    },

                            {
                                'label': 'Corrección',
                                'data': {'2': 3},
                                'backgroundColor': [
                                    'rgba(255, 206, 86, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(255, 206, 86, 1)',
                                ],
                                'borderWidth': 1
                            },

                            {
                                'label': 'No procesado',
                                'data': {'1': 4},
                                'backgroundColor': [
                                    'rgba(75, 192, 192, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(75, 192, 192, 1)',
                                ],
                                'borderWidth': 1
                            }
                        ]
                    }
        },
        'grafico4':{
            'data':{
                    'labels': ['Juan Espinosa','Camilo Sarmiento', 'Flor Cardenas', 'Maria Rivera'],
                    'datasets':
                    [
                        {
                            # 'axis': 'y',
                            'label': 'OK Procesado',
                             'data':  [6, 5, 6,3],
                                'backgroundColor': [
                                     'rgba(54, 162, 235, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(54, 162, 235, 1)',
                                ],
                                'borderWidth': 1
                        },
                        {
                            # 'axis': 'y',
                            'label': 'Pendiente',
                            'data': [ 3, 3, 3, 2],
                                'backgroundColor': [
                                    'rgba(255, 99, 132, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(255, 99, 132, 1)',
                                ],
                                'borderWidth': 1
                        },
                        {
                            # 'axis': 'y',
                            'label': 'No procesado',
                            'data': [ 1,1,6,1 ],
                                'backgroundColor': [
                                    'rgba(255, 206, 86, 0.2)',
                                ],
                                'borderColor': [
                                    'rgba(255, 206, 86, 1)',
                                ],
                                'borderWidth': 1
                        },

                    ]

                    }
                    }}

